#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Создание примера шаблона Excel для Audit Processor
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Создание новой книги
wb = Workbook()
ws = wb.active
ws.title = "Аудит"

# Заголовки колонок (примерная структура таблицы аудита)
headers = [
    "№",
    "Дата документа",
    "Номер документа",
    "Тип документа",
    "Название/Описание",
    "Сумма",
    "Контрагент",
    "Ответственное лицо",
    "Статус",
    "Выявленные несоответствия",
    "Рекомендации",
    "Примечания"
]

# Стилизация заголовков
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Установка заголовков
for col_num, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = border

# Установка ширины колонок
column_widths = [5, 12, 15, 15, 30, 12, 20, 20, 12, 35, 35, 25]
for col_num, width in enumerate(column_widths, start=1):
    ws.column_dimensions[chr(64 + col_num)].width = width

# Установка высоты строки заголовка
ws.row_dimensions[1].height = 30

# Добавление нескольких примеров строк (опционально)
examples = [
    ["", "01.11.2025", "ДОК-001", "Договор", "Договор поставки оборудования", "150000", "ООО «Поставщик»", "Иванов И.И.", "Проверен", "", "", ""],
    ["", "05.11.2025", "АКТ-015", "Акт приемки", "Акт приемки выполненных работ", "85000", "ООО «Подрядчик»", "Петров П.П.", "На проверке", "", "", ""],
]

for row_num, example in enumerate(examples, start=2):
    for col_num, value in enumerate(example, start=1):
        cell = ws.cell(row=row_num, column=col_num, value=value)
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)

# Сохранение файла
output_path = "audit_template.xlsx"
wb.save(output_path)

print(f"✅ Шаблон Excel создан: {output_path}")
print(f"📊 Колонок: {len(headers)}")
print(f"📝 Заголовки: {', '.join(headers[:4])}...")
